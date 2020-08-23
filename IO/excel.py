import openpyxl
import pandas as pd
#創建sheet
#更名
#將資料集填入sheet(集合的型態)
def StuffDataToExcel(datas,sheets,outputName):#'Data.xlsx'
    wb = Workbook()
    order =0
    for item in sheets:
        order =order+1
        wb.create_sheet(item, index=order)
        
    stuffSheetNames = map(StuffWhichSheet,datas)
    for x, y in zip(stuffSheetNames, datas):
        wb[x].append(y)    
    wb.save(outputName)

#單一筆
def StuffATableToExcel(data,sheet,outputName):
    df=pd.DataFrame(data=data)    
    wb=openpyxl.load_workbook(outputName)
    writer=pd.ExcelWriter(outputName,engine="openpyxl")
    df.to_excel(writer,sheet_name='永慶')
    writer.save()
    writer.close()
    #没有下面这个语句的话excel表将完全被覆盖
    #writer.book=wb
    #wb._active_sheet_index=1#修改作用中的sheet
    #active_sheet = wb.active


def getRowDataFromExcel(pathAddFile,sheet,index):
    workbook = openpyxl.load_workbook(pathAddFile)
    worksheet = workbook.get_sheet_by_name(sheet)
    rowData = [item.value for item in list(worksheet.rows)[index]]
    return rowData


def getUrlFromExcel(pathAddFile,sheet,index,title):
    workbook = openpyxl.load_workbook(pathAddFile)
    worksheet = workbook.get_sheet_by_name(sheet)
    urls = [item.value for item in list(worksheet.columns)[index]]
    urls.remove(title)
    return urls

#依據資料集決定sheet的名稱
def StuffWhichSheet(data):
   if '標示部' in data[0] != -1:
      return '標示部'
   elif '所有權部' in data[0]:
      return '所有權部'
   elif '權利部' in data[0]:
      return '權利部'



 
